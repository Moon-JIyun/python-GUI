from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import time
import tempfile
from PIL import Image as PILImage
from io import BytesIO

# 이미지 파일이 있는 폴더 경로
result_folder = './Result'
samples = [sample for sample in os.listdir(result_folder) if not sample.startswith('.DS_Store')]
print(samples)

# 엑셀 파일 생성
workbook = Workbook()
start_time = time.time()
print(start_time)
for sample in samples:
    direction_folder = os.path.join(result_folder, sample)
    directions = [direction for direction in os.listdir(direction_folder) if not direction.startswith('.DS_Store')]
    print(directions)
    directions = sorted(directions)
    print(directions)
    # 이미지 파일 목록 읽기
    new_sheet = workbook.create_sheet(title= sample)

    # 생성된 시트를 활성화
    workbook.active = new_sheet
    sheet = workbook.active
    print(sheet.title)

    column = 3
    for direction in directions:
        image_folder = os.path.join(direction_folder, direction)
        image_files = [image for image in os.listdir(image_folder) if not image.startswith('.DS_Store')]
        print(image_files)
        if image_files[0][0] == "-":
            image_files = sorted(image_files, key=lambda x: int(x.split('.')[0]), reverse= True)
            print("-degree::", image_files)
        else:
            image_files = sorted(image_files, key=lambda x: int(x.split('.')[0]))
            print("+degree:", image_files)

        if direction[-1] == "-":
            row = 4
            # sheet.merge_cells(start_row=row - 2, start_column=column - 1, end_row=row - 2, end_column=column)
            # sheet.cell(row=row - 2, column=column - 1).value = direction
            merge_cell_start = sheet.cell(row=row - 2, column=column - 1)
            merge_cell_end = sheet.cell(row=row - 2, column=column)
            merge_cell_range = f'{merge_cell_start.coordinate}:{merge_cell_end.coordinate}'
            sheet.merge_cells(merge_cell_range)
            merge_cell_start.value = direction
            merge_cell_start.alignment = Alignment(horizontal='center', vertical='center')

        else:
            # row = 3
            # sheet.merge_cells(start_row = row-1, start_column = column-1, end_row = row-1, end_column = column)
            # sheet.cell(row=row-1, column = column-1).value = direction
            row = 3
            merge_cell_start = sheet.cell(row=row - 1, column=column - 1)
            merge_cell_end = sheet.cell(row=row - 1, column=column)
            merge_cell_range = f'{merge_cell_start.coordinate}:{merge_cell_end.coordinate}'
            sheet.merge_cells(merge_cell_range)
            merge_cell_start.value = direction
            merge_cell_start.alignment = Alignment(horizontal='center', vertical='center')

        for image_file in image_files:
            image_path = os.path.join(image_folder, image_file)
            print("imgae_path:", image_path)

            image = PILImage.open(image_path)
            img_width, img_height = int(image.width*0.1), int(image.height*0.1)
            image.thumbnail((img_width, img_height))

            img_bytes = BytesIO()
            image.save(img_bytes, format='JPEG')

            # temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            # temp_file_path = temp_file.name
            # image.save(temp_file_path)
            # temp_file.close()

            # resized_img = img.resize((img_width, img_height))
            # resized_img.save(image_path)


            # img_width, img_height = int(img.width*0.1), int(img.height*0.1)
            # print("before:", img.width, img.height)
            # img.width, img.height = img_width, img_height
            # print("after:", img.width, img.height)
            # 이미지 삽입 셀의 크기 조정
            sheet.column_dimensions[get_column_letter(column)].width = img_width * 0.14
            sheet.row_dimensions[row].height = img_height * 0.75

            cell = sheet.cell(row=row, column=column)
            # 이미지 삽입

            img = Image(img_bytes)
            print(img)
            sheet.add_image(img, f'{get_column_letter(column)}{row}')
            # 이미지 파일명 삽입
            file_name_cell = sheet.cell(row=row, column=column -1)
            file_name_cell.value = image_file[:-4]

            # 다음 셀로 이동
            row += 1

            # os.remove((temp_file_path))
            # os.remove(resized_img)
        column += 2
        # 엑셀 파일 저장
end_time = time.time()
print(f"{end_time - start_time:.5f} sec")
workbook.save('./Result/output_1.xlsx')